"""Excel export via openpyxl.

Template structure (DGC-curve-calculator.xlsx) — confirmed from file inspection:

  4 sheets in 2 pairs, one pair per dot shape:
    Pair 0 (shapes[0]): Sheet1 = weight[0] (120#), Sheet2 = weight[1] (150#) + weight[2] (200#)
    Pair 1 (shapes[1]): Sheet3 = weight[0],        Sheet4 = weight[1] + weight[2]

  Single-weight sheet layout (Sheet1 / Sheet3):
    A1      : job title  "{customer} {stock}"
    I1      : date
    B4:E17  : CMYK step readings for weight[0], steps 100→1  (14 rows × 4 cols)
              or B4:E19 for 16-step jobs (adds 0.8%, 0.4%)
    A20     : weight[0].label  (row = 4 + num_steps + 2)
    I20     : job.crs

  Dual-weight sheet layout (Sheet2 / Sheet4):
    A1      : job title
    I1      : date
    B4:E17  : CMYK step readings for weight[1]
    A20     : weight[1].label
    I20     : job.crs
    B27:E40 : CMYK step readings for weight[2]
    A43     : weight[2].label
    I43     : job.crs

  Note: rows 18–19 (0.8%, 0.4% steps) and rows 41–42 are filled for 16-step jobs.

  Bug fix: In Sheet2/Sheet4 the second table's F-column formulas (F27:F40) reference
  wrong rows due to a copy-paste error in the original template.  We rewrite them with
  the correct row references when exporting.

  The density readings (2.11, 1.80, etc.) are written to the Illustrator PDF only;
  the Excel template has no density row.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.cell import MergedCell
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

from core.models import JobConfig, ShapeData


# Row where step data starts (first table, both sheet types)
_STEP_START_ROW_T1 = 4
# Gap (in rows) between the T1 label row and the T2 step-start row
_GAP_T1_TO_T2 = 7
# CMYK data columns: B=2, C=3, D=4, E=5
_DATA_COLS = [2, 3, 4, 5]


def _row_constants(num_steps: int) -> tuple[int, int, int]:
    """Return (label_t1, step_start_t2, label_t2) for a given step count.

    Derivation (verified against both bundled templates):
      label_t1      = STEP_START_T1 + num_steps          (label immediately follows last step)
      step_start_t2 = label_t1 + GAP_T1_TO_T2
      label_t2      = step_start_t2 + num_steps
    """
    label_t1 = _STEP_START_ROW_T1 + num_steps
    step_start_t2 = label_t1 + _GAP_T1_TO_T2
    label_t2 = step_start_t2 + num_steps
    return label_t1, step_start_t2, label_t2


def _bundled_template(num_steps: int) -> Path:
    """Return the path to the bundled template matching the step count."""
    base = Path(sys._MEIPASS) if getattr(sys, "frozen", False) else Path(__file__).parent.parent
    name = "template_extended.xlsx" if num_steps > 14 else "template_standard.xlsx"
    return base / "assets" / name


def export_excel(job: JobConfig, output_path: str) -> None:
    """Fill the Excel template from job data and save to output_path."""
    num_steps = len(job.step_labels)

    # Use explicit override path if set, otherwise auto-select bundled template
    if job.template_xlsx_path:
        template_path = Path(job.template_xlsx_path)
    else:
        template_path = _bundled_template(num_steps)

    if not template_path.is_file():
        raise FileNotFoundError(f"Excel template not found: {template_path}")

    wb = openpyxl.load_workbook(str(template_path))
    label_t1, step_start_t2, label_t2 = _row_constants(num_steps)

    title = " ".join(filter(None, [
        job.customer, job.print_type, job.stock_desc, job.finish
    ]))
    dot_shape = " ".join(filter(None, [job.dot_shape_type, job.dot_shape_number]))

    for shape_idx, shape in enumerate(job.shapes):
        # Each shape occupies a pair of sheets (indices 2*shape_idx and 2*shape_idx + 1)
        single_idx = shape_idx * 2      # Sheet1, Sheet3, Sheet5…
        dual_idx   = shape_idx * 2 + 1  # Sheet2, Sheet4, Sheet6…

        # If the template doesn't have enough sheets, add new ones by copying
        while len(wb.sheetnames) <= dual_idx:
            _append_sheet_pair(wb, num_steps)

        ws_single = wb[wb.sheetnames[single_idx]]
        ws_dual   = wb[wb.sheetnames[dual_idx]]

        # --- Single sheet: weight[0] ---
        _write_metadata(ws_single, title, job.date)
        if len(shape.weights) > 0:
            w0 = shape.weights[0]
            _write_steps(ws_single, w0.steps, _STEP_START_ROW_T1, num_steps)
            _write_cell(ws_single, label_t1, 1, w0.label)
            _write_cell(ws_single, label_t1, 9, dot_shape)

        # --- Dual sheet: weight[1] (first table) ---
        _write_metadata(ws_dual, title, job.date)
        if len(shape.weights) > 1:
            w1 = shape.weights[1]
            _write_steps(ws_dual, w1.steps, _STEP_START_ROW_T1, num_steps)
            _write_cell(ws_dual, label_t1, 1, w1.label)
            _write_cell(ws_dual, label_t1, 9, dot_shape)

        # --- Dual sheet: weight[2] (second table) ---
        if len(shape.weights) > 2:
            w2 = shape.weights[2]
            _write_steps(ws_dual, w2.steps, step_start_t2, num_steps)
            _write_cell(ws_dual, label_t2, 1, w2.label)
            _write_cell(ws_dual, label_t2, 9, dot_shape)
            _fix_second_table_formulas(ws_dual, step_start_t2, num_steps)

    # Apply A4 page setup to all worksheets
    for ws in wb.worksheets:
        ws.page_setup.paperSize = 9           # 9 = A4
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0         # auto height
        # fitToPage must be set via sheet_properties to avoid NoneType error
        # on worksheets loaded from existing files
        if ws.sheet_properties is None:
            ws.sheet_properties = WorksheetProperties()
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_area = f"A1:I{ws.max_row}"
        ws.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.75, bottom=0.75,
            header=0.3, footer=0.3,
        )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_cell(ws, row: int, col: int, value) -> None:
    """Write value to a cell, resolving to the top-left cell if it is part of a merge."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col, value=value)
                return
    else:
        cell.value = value


def _write_metadata(ws, title: str, date: str) -> None:
    _write_cell(ws, 1, 1, title)   # A1
    _write_cell(ws, 1, 9, date)    # I1


def _write_steps(ws, steps: list[list[float]], start_row: int, num_steps: int) -> None:
    """Write num_steps rows of CMYK step readings into columns B–E starting at start_row."""
    for ri in range(num_steps):
        if ri >= len(steps):
            break
        row_data = steps[ri]
        excel_row = start_row + ri
        for ci, col in enumerate(_DATA_COLS):
            value = row_data[ci] if ci < len(row_data) else None
            # Write blank rather than 0 for unset values so formulas read as empty
            _write_cell(ws, excel_row, col, value if value else None)


def _fix_second_table_formulas(ws, start_row: int, num_steps: int) -> None:
    """Correct the Average% formulas in the second table's F column.

    The original template has a copy-paste error where F27 references B25:E25
    instead of B27:E27 (offset by 2 rows).  We overwrite F{start_row}:F{end_row}
    with correct =SUM(B{r}:E{r})/4 formulas.
    """
    for i in range(num_steps):
        r = start_row + i
        ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})/4")


def _append_sheet_pair(wb: openpyxl.Workbook, num_steps: int) -> None:
    """Add a blank single+dual sheet pair by copying Sheet1/Sheet2 structure."""
    existing = wb.sheetnames
    src_single = wb[existing[0]]
    src_dual   = wb[existing[1]]

    new_single = wb.copy_worksheet(src_single)
    new_dual   = wb.copy_worksheet(src_dual)

    _, step_start_t2, _ = _row_constants(num_steps)
    t1_clear_end = _STEP_START_ROW_T1 + num_steps + 1
    t2_clear_end = step_start_t2 + num_steps + 1

    for ws in (new_single, new_dual):
        for row in range(_STEP_START_ROW_T1, t1_clear_end):
            for col in range(2, 6):
                ws.cell(row=row, column=col, value=None)
    for row in range(step_start_t2, t2_clear_end):
        for col in range(2, 6):
            new_dual.cell(row=row, column=col, value=None)
